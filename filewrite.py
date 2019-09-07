import openpyxl
import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows

def filewrite(datefw, pairs, bets):
        
        # Excel tiedoston muokkaus
        print('Opening existing file...')
        try:
        # Trying to open MLB.xlsx
                wb = openpyxl.load_workbook('MLB.xlsx')
        except IOError:
                print("File does not exist, I will make one for you.")
                wb = openpyxl.Workbook()
                sheet = wb.get_sheet_by_name('Sheet')
                sheet.title = 'Start'
                sheet['A2'] = 'Python program saves betting tips to this excel file, tips are always saved to new sheets which are named by date.'
                wb.save('MLB.xlsx')


        # Creating new sheet
        sheet = wb.create_sheet(index = 0, title = datefw)

        ws = wb.active

        df = pd.DataFrame({'Todays games:'+ ' ' + datefw: pairs, 'O/U bets: ' : bets})
        
        for r in dataframe_to_rows(df, index=False):
                ws.append(r)

        wb.save('MLB.xlsx')

        print("Done!")